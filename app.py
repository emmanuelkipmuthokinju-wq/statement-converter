import streamlit as st
import io, re
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

st.set_page_config(page_title="Statement PDF → Excel", page_icon="📄", layout="centered")
st.markdown("""
<style>
.hero{background:linear-gradient(135deg,#1e3a5f,#2563eb);border-radius:16px;
      padding:36px 32px 28px;margin-bottom:28px;color:white;text-align:center}
.hero h1{font-size:2rem;margin:0 0 8px;font-weight:700}
.hero p{font-size:1rem;opacity:.85;margin:0}
.badge{display:inline-block;background:rgba(255,255,255,.18);border-radius:20px;
       padding:3px 14px;font-size:.8rem;margin:10px 4px 0}
.ok{background:#f0fdf4;border:1px solid #86efac;border-radius:10px;
    padding:16px 20px;color:#166534;font-weight:500;margin-top:12px}
.stDownloadButton>button{background:#2563eb!important;color:white!important;
    border-radius:8px!important;padding:10px 24px!important;font-size:1rem!important;
    font-weight:600!important;width:100%!important;border:none!important}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero">
  <h1>📄 Statement → Excel</h1>
  <p>Upload a customer statement PDF and download a clean, formatted Excel file instantly.</p>
  <span class="badge">✅ Savannah Cement</span>
  <span class="badge">✅ National Cement</span>
  <span class="badge">✅ Mombasa Cement</span>
  <span class="badge">✅ Karsan Ramji</span>
  <span class="badge">🔒 Files never stored</span>
</div>
""", unsafe_allow_html=True)

with st.expander("📋 Supported formats"):
    st.markdown("""
| Supplier | Key Columns |
|---|---|
| 🟠 **Savannah Cement** | Doc No · Ref/Vehicle · Date · Due Date · Details · Debit · Credit · Balance |
| 🔵 **National Cement** | Date · Description · Vehicle · LPO · Qty · Invoice · CU Invoice · Location · Debit · Credit · Balance + Cheques sheet |
| 🟢 **Mombasa Cement** | Doc Date · Reference · Cheque No · Description · Amount · Balance |
| 🟣 **Karsan Ramji (Ndovu)** | Date · Doc No · BP Ref · CU Invoice · Line Memo · Debit · Credit · Balance |
    """)

# ── Colours & helpers ──────────────────────────────────────────────────────────
C = dict(savannah_orange="E8540A", national_blue="1565C0", mombasa_green="1B5E20",
         karsan_purple="4A148C", header_dark="2C3E50", col_hdr_fg="FFFFFF",
         alt_row="F2F6FA", credit_row="E8F5E9", receipt_row="FFF3E0",
         info_bg="ECF0F1", debit_fg="1565C0", credit_fg="2E7D32",
         neg_bal="C62828", border="BDBDBD")
MFMT = '#,##0.00;(#,##0.00);"-"'

def thin(sides="all"):
    s = Side(style="thin", color=C["border"])
    kw = dict(left=s,right=s,top=s,bottom=s) if sides=="all" else dict(bottom=s)
    return Border(**kw)

def pn(s):
    if s is None: return None
    try: return float(str(s).strip().replace(",","").replace("-","").strip()) * (-1 if str(s).strip().endswith("-") else 1)
    except: return None

def wc(ws, r, c, v=None, bold=False, fg="000000", bg=None, align="left",
       fmt=None, bdr=None, sz=9, ind=0):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", size=sz, bold=bold, color=fg)
    if bg: cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", indent=ind)
    if fmt: cell.number_format = fmt
    if bdr: cell.border = bdr
    return cell

def banner(ws, row, cols, text, bg):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    wc(ws, row, 1, text, bold=True, fg=C["col_hdr_fg"], bg=bg, align="center", sz=14)
    ws.row_dimensions[row].height = 30

def col_headers(ws, row, headers, bg):
    for ci, h in enumerate(headers, 1):
        wc(ws, row, ci, h, bold=True, fg=C["col_hdr_fg"], bg=bg, align="center", bdr=thin())
    ws.row_dimensions[row].height = 22

def totals_row(ws, row, merge_end, label_text, data_cols, bg):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_end)
    wc(ws, row, 1, label_text, bold=True, fg=C["col_hdr_fg"], bg=bg, align="right", bdr=thin(), ind=1)
    for ci, col_letter in data_cols:
        c = ws.cell(row=row, column=ci, value=f"=SUM({col_letter}{row-len(data_cols[0])+1 if False else ''}")
        # just use formula directly
    ws.row_dimensions[row].height = 18

# ── FORMAT DETECTION ───────────────────────────────────────────────────────────
def detect(lines):
    t = "\n".join(lines[:50])
    if "MOMBASA CEMENT" in t.upper():                                  return "mombasa"
    if "Karsan Ramji" in t or "NDOVU" in t.upper() or "BP REF" in t:  return "karsan"
    if "NATIONAL CEMENT" in t.upper() or "CU Invoice Number" in t:    return "national"
    if "Savannah Cement" in t or "A/R Invoices" in t:                 return "savannah"
    if "LPO No" in t or "Cheque no" in t:                             return "national"
    if "Due Date" in t and "Document" in t:                           return "savannah"
    if "AR Invoice" in t and "BP REF" in t:                           return "karsan"
    return "unknown"

# ══════════════════════════════════════════════════════════════════════════════
# MOMBASA CEMENT PARSER
# ══════════════════════════════════════════════════════════════════════════════
DATE_MCL = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
MONEY_RE  = re.compile(r"^-?[\d,]+\.\d{2}-?$")

def parse_mombasa(lines):
    hdr = {"supplier": "Mombasa Cement Limited", "account": "", "account_no": "",
           "currency": "KES", "period": ""}
    for ln in lines[:15]:
        m = re.search(r"ACCOUNT:\s*(.+?)\s{2,}", ln)
        if m: hdr["account"] = m.group(1).strip()
        m = re.search(r"ACCOUNT NO\s+(\S+)", ln)
        if m: hdr["account_no"] = m.group(1)
        m = re.search(r"CURRENCY\s+(\S+)", ln)
        if m: hdr["currency"] = m.group(1)
    for ln in lines:
        m = re.search(r"PERIOD\s+([\d.]+)\s+TO\s+([\d.]+)", ln)
        if m: hdr["period"] = f"{m.group(1)} to {m.group(2)}"; break

    txs = []
    skip = {"DOC. DATE","OUR","REFERENCE","CHEQUE NO","DESCRIPTION","AMOUNT",
            "BALANCE","MOMBASA CEMENT","P. O. BOX","MOMBASA,KENYA","Tel:",
            "Mobile:","Email:","STATEMENT","ACCOUNT:","NAIROBI","KENYA","PAGE",
            "NET TOTAL","CLOSING BALANCE","AGEING","UNALLOCATED","AMOUNT DUE",
            "1 TO","31 TO","61 TO","91 TO","OVER","We consider","writing within",
            "Please note","TRANSACTION DURING","OPENING BALANCE"}

    section = "opening"
    for ln in lines:
        ln = ln.strip()
        if "TRANSACTION DURING PERIOD" in ln: section = "current"; continue
        if "CLOSING BALANCE" in ln:
            p = ln.split()
            amts = [x for x in p if MONEY_RE.match(x)]
            if amts: hdr["closing_balance"] = pn(amts[-1])
            continue
        if any(ln.startswith(s) or s in ln for s in skip if len(s) > 3): continue

        parts = ln.split()
        if not parts or not DATE_MCL.match(parts[0]): continue

        date = parts[0]; rest = parts[1:]
        # find money amounts at end
        amts_idx = []
        for j in range(len(rest)-1, -1, -1):
            if MONEY_RE.match(rest[j]): amts_idx.insert(0, j)
            elif amts_idx: break

        if not amts_idx: continue
        first_amt = amts_idx[0]
        desc_parts = rest[:first_amt]
        amt_parts  = rest[first_amt:]

        ref_no = desc_parts[0] if desc_parts else ""
        cheque = ""
        desc   = ""
        # if 3+ desc parts, second might be cheque if alphanumeric, rest is description
        if len(desc_parts) >= 3:
            if re.match(r'^[A-Z0-9]+$', desc_parts[1]) and not MONEY_RE.match(desc_parts[1]):
                cheque = desc_parts[1]; desc = " ".join(desc_parts[2:])
            else:
                desc = " ".join(desc_parts[1:])
        elif len(desc_parts) == 2:
            desc = desc_parts[1]

        amount = pn(amt_parts[0]) if amt_parts else None
        balance= pn(amt_parts[1]) if len(amt_parts) > 1 else None

        is_receipt = "Receipt" in desc or "receipt" in desc
        txs.append({"date": date, "ref_no": ref_no, "cheque_no": cheque,
                    "description": desc, "amount": amount, "balance": balance,
                    "section": section, "is_receipt": is_receipt})
    return hdr, txs

def excel_mombasa(hdr, txs):
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = "Statement"; ws.sheet_properties.tabColor = C["mombasa_green"]
    for i, w in enumerate([14, 14, 16, 14, 28, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    banner(ws, 1, 7, "MOMBASA CEMENT LIMITED  |  CUSTOMER STATEMENT", C["mombasa_green"])

    # Info block
    lf = Font(name="Arial", bold=True, size=9)
    vf = Font(name="Arial", size=9)
    lb = PatternFill("solid", start_color=C["info_bg"])
    al = Alignment(horizontal="left", vertical="center", indent=1)
    info = [("Account", hdr.get("account",""), "Account No", hdr.get("account_no","")),
            ("Currency", hdr.get("currency","KES"), "Period", hdr.get("period",""))]
    for r, (l1,v1,l2,v2) in enumerate(info, start=2):
        for ci,lbl,val in [(1,l1,v1),(5,l2,v2)]:
            ws.merge_cells(start_row=r,start_column=ci,end_row=r,end_column=ci+1)
            ws.merge_cells(start_row=r,start_column=ci+2,end_row=r,end_column=ci+2)
            lc=ws.cell(row=r,column=ci,value=lbl); lc.font=lf; lc.fill=lb; lc.alignment=al
            vc=ws.cell(row=r,column=ci+2,value=val); vc.font=vf; vc.alignment=al
        ws.row_dimensions[r].height=16

    H = 4
    col_headers(ws, H, ["Doc Date","Our Reference","Cheque No","Section",
                         "Description","Amount (KES)","Balance (KES)"], C["mombasa_green"])

    D = H + 1
    for ri, tx in enumerate(txs, start=D):
        is_rec = tx["is_receipt"]
        is_open= tx["section"] == "opening"
        bg = C["receipt_row"] if is_rec else (C["alt_row"] if (ri-D)%2==1 else None)
        sec_label = "Opening Bal." if is_open else "Current Period"

        vals   = [tx["date"], tx["ref_no"], tx["cheque_no"], sec_label,
                  tx["description"], tx["amount"], tx["balance"]]
        aligns = ["center","center","center","center","left","right","right"]
        fmts   = [None,None,None,None,None,MFMT,MFMT]
        for ci,(v,a,f) in enumerate(zip(vals,aligns,fmts),1):
            fg="000000"
            if ci==6 and v is not None: fg=C["credit_fg"] if is_rec else C["debit_fg"]
            if ci==7 and v is not None: fg=C["neg_bal"] if v<0 else C["credit_fg"]
            wc(ws,ri,ci,v,align=a,fmt=f,fg=fg,bg=bg,bdr=thin("bottom"))
        ws.row_dimensions[ri].height=15

    T = D + len(txs)
    ws.merge_cells(f"A{T}:E{T}")
    wc(ws,T,1,"TOTAL / CLOSING BALANCE",bold=True,fg=C["col_hdr_fg"],
       bg=C["mombasa_green"],align="right",bdr=thin(),ind=1)
    for ci,col in [(6,"F"),(7,"G")]:
        c=ws.cell(row=T,column=ci,value=f"=SUM({col}{D}:{col}{T-1})")
        c.font=Font(name="Arial",bold=True,size=9,color=C["col_hdr_fg"])
        c.fill=PatternFill("solid",start_color=C["mombasa_green"])
        c.number_format=MFMT; c.alignment=Alignment(horizontal="right",vertical="center")
        c.border=thin()
    ws.row_dimensions[T].height=18

    ws.freeze_panes=f"A{D}"; ws.auto_filter.ref=f"A{H}:G{T-1}"
    ws.page_setup.orientation="landscape"; ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ══════════════════════════════════════════════════════════════════════════════
# KARSAN RAMJI PARSER
# ══════════════════════════════════════════════════════════════════════════════
DATE_KRS = re.compile(r"^\d{2}/\d{2}/\d{4}$")
MONEY_KRS= re.compile(r"^-?[\d,]+\.\d{2}$")

def parse_karsan(lines):
    hdr = {"supplier":"Karsan Ramji and Sons Ltd","customer":"","customer_code":"",
           "phone":"","email":"","period":""}
    for i, ln in enumerate(lines[:10]):
        m = re.search(r"(N\d+)\s+(.+)", ln)
        if m: hdr["customer_code"]=m.group(1); hdr["customer"]=m.group(2).strip()
        m = re.search(r"Phone#?:\s*(\S+)", ln)
        if m: hdr["phone"]=m.group(1)
        m = re.search(r"Email:\s*(\S+)", ln)
        if m: hdr["email"]=m.group(1)
        m = re.search(r"From\s+([\d/]+).+To\s+([\d/]+)", ln)
        if m: hdr["period"]=f"{m.group(1)} to {m.group(2)}"

    txs = []
    skip_kw = {"DATE","DOC. NO.","BP REF.","CU INVOICE","LINE MEMO","DEBIT","CREDIT",
               "BALANCE","Customer Statement","Document Date","Karsan","Phone","Address",
               "Email","Current Balance","Total","Aging","E & O","Page"}

    for ln in lines:
        ln = ln.strip()
        if any(k in ln for k in skip_kw): continue

        # Balance B/F line
        if "Balance B/F" in ln:
            p = ln.split()
            amts = [x for x in p if MONEY_KRS.match(x)]
            bal = pn(amts[-1]) if amts else None
            txs.append({"date":"--","doc_no":"--","bp_ref":"--","cu_invoice":"--",
                        "line_memo":"Balance B/F","debit":None,"credit":None,
                        "balance":bal,"is_bf":True})
            continue

        parts = ln.split()
        if not parts or not DATE_KRS.match(parts[0]): continue

        date=parts[0]; rest=parts[1:]
        # amounts are last 3 numbers
        amts=[]; nonamts=[]
        for x in reversed(rest):
            if MONEY_KRS.match(x) and len(amts)<3: amts.insert(0,x)
            else: nonamts.insert(0,x)

        dbt = pn(amts[0]) if len(amts)>=3 else None
        crd = pn(amts[1]) if len(amts)>=3 else None
        bal = pn(amts[2]) if len(amts)>=3 else (pn(amts[-1]) if amts else None)

        na=nonamts
        doc_no  = na[0] if len(na)>0 else ""
        bp_ref  = na[1] if len(na)>1 else ""
        cu_inv  = na[2] if len(na)>2 else ""
        memo    = " ".join(na[3:]) if len(na)>3 else ""

        txs.append({"date":date,"doc_no":doc_no,"bp_ref":bp_ref,"cu_invoice":cu_inv,
                    "line_memo":memo,"debit":dbt,"credit":crd,"balance":bal,"is_bf":False})
    return hdr, txs

def excel_karsan(hdr, txs):
    wb=openpyxl.Workbook(); ws=wb.active
    ws.title="Statement"; ws.sheet_properties.tabColor=C["karsan_purple"]
    for i,w in enumerate([12,14,12,22,14,16,16,18],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    banner(ws,1,8,"KARSAN RAMJI AND SONS LTD (NDOVU CEMENT)  |  CUSTOMER STATEMENT",C["karsan_purple"])

    lf=Font(name="Arial",bold=True,size=9); vf=Font(name="Arial",size=9)
    lb=PatternFill("solid",start_color=C["info_bg"])
    al=Alignment(horizontal="left",vertical="center",indent=1)
    info=[("Customer",hdr.get("customer",""),"Code",hdr.get("customer_code","")),
          ("Period",hdr.get("period",""),"Phone",hdr.get("phone",""))]
    for r,(l1,v1,l2,v2) in enumerate(info,start=2):
        for ci,lbl,val in [(1,l1,v1),(5,l2,v2)]:
            ws.merge_cells(start_row=r,start_column=ci,end_row=r,end_column=ci+1)
            ws.merge_cells(start_row=r,start_column=ci+2,end_row=r,end_column=ci+2)
            lc=ws.cell(row=r,column=ci,value=lbl); lc.font=lf; lc.fill=lb; lc.alignment=al
            vc=ws.cell(row=r,column=ci+2,value=val); vc.font=vf; vc.alignment=al
        ws.row_dimensions[r].height=16

    H=4
    col_headers(ws,H,["Date","Doc No.","BP Ref No.","CU Invoice No.",
                       "Line Memo","Debit (KES)","Credit (KES)","Balance (KES)"],C["karsan_purple"])

    D=H+1
    for ri,tx in enumerate(txs,start=D):
        is_bf=tx.get("is_bf",False)
        bg=C["info_bg"] if is_bf else (C["alt_row"] if (ri-D)%2==1 else None)
        vals=[tx["date"],tx["doc_no"],tx["bp_ref"],tx["cu_invoice"],
              tx["line_memo"],tx["debit"],tx["credit"],tx["balance"]]
        aligns=["center","center","center","center","left","right","right","right"]
        fmts=[None,None,None,None,None,MFMT,MFMT,MFMT]
        for ci,(v,a,f) in enumerate(zip(vals,aligns,fmts),1):
            fg="000000"
            if ci==6 and v: fg=C["debit_fg"]
            if ci==7 and v: fg=C["credit_fg"]
            if ci==8 and v is not None: fg=C["neg_bal"] if v<0 else C["credit_fg"]
            wc(ws,ri,ci,v,align=a,fmt=f,fg=fg,bg=bg,bdr=thin("bottom"))
        ws.row_dimensions[ri].height=15

    T=D+len(txs)
    ws.merge_cells(f"A{T}:E{T}")
    wc(ws,T,1,"CURRENT BALANCE",bold=True,fg=C["col_hdr_fg"],bg=C["karsan_purple"],
       align="right",bdr=thin(),ind=1)
    for ci,col in [(6,"F"),(7,"G"),(8,"H")]:
        c=ws.cell(row=T,column=ci,value=f"=SUM({col}{D}:{col}{T-1})")
        c.font=Font(name="Arial",bold=True,size=9,color=C["col_hdr_fg"])
        c.fill=PatternFill("solid",start_color=C["karsan_purple"])
        c.number_format=MFMT; c.alignment=Alignment(horizontal="right",vertical="center")
        c.border=thin()
    ws.row_dimensions[T].height=18
    ws.freeze_panes=f"A{D}"; ws.auto_filter.ref=f"A{H}:H{T-1}"
    ws.page_setup.orientation="landscape"; ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ══════════════════════════════════════════════════════════════════════════════
# NATIONAL CEMENT PARSER  (precise column-aware parser)
# Columns: Date | Description | Cheque no | Vehicle No. | LPO No | Quantity |
#          Invoice No. | CU Invoice Number | Location | Curr. | Debit | Credit | Balance
# ══════════════════════════════════════════════════════════════════════════════
DR4     = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
MONEY_N = re.compile(r"^[\d,]+\.\d{2}$")
QTY_N   = re.compile(r"^\d+\.\d{3}$")
LPOSUF  = re.compile(r"^\d{7,8}$")   # LPO suffix split to next line e.g. 06005663

def parse_national(lines):
    hdr = {}
    for ln in lines[:20]:
        m = re.search(r"Statement As At:\s*([\d.]+)", ln)
        if m: hdr["statement_date"] = m.group(1)
        m = re.search(r"PIN No:\s*(\S+)", ln)
        if m: hdr["pin"] = m.group(1)
        if "NATIONAL CEMENT" in ln.upper(): hdr["supplier"] = "National Cement Company Ltd"
        if "New Muthokinju" in ln: hdr["customer"] = "New Muthokinju Hardware Ltd"
    hdr.setdefault("supplier", "National Cement Company Ltd")
    hdr.setdefault("customer", "New Muthokinju Hardware Ltd")

    txs = []; chqs = []; in_chq = False
    i = 0
    while i < len(lines):
        ln = lines[i].strip()

        if "Cheques on Hand" in ln: in_chq = True; i+=1; continue
        if "Cheque No" in ln and "Due On" in ln: i+=1; continue
        if "Total/Closing" in ln or "Ageing" in ln or "Total" == ln.split()[0] if ln.split() else False:
            i+=1; continue

        if in_chq:
            p = ln.split()
            if len(p) == 3 and DR4.match(p[1]):
                try: chqs.append({"cheque_no":p[0],"due_on":p[1],"amount":float(p[2].replace(",",""))})
                except: pass
            i+=1; continue

        parts = ln.split()
        if not parts or not DR4.match(parts[0]): i+=1; continue

        date = parts[0]
        rest = list(parts[1:])

        # Consume LPO suffix on next line (pure 7-8 digit number like 06005663)
        lpo_suffix = ""
        if i+1 < len(lines):
            nxt = lines[i+1].strip()
            if LPOSUF.match(nxt):
                lpo_suffix = nxt
                i += 1

        # Balance B/F line
        if rest and rest[0] == "Balance":
            dr_cr = rest[-1] if rest[-1] in ("DR","CR") else ""
            nums = [x for x in rest if MONEY_N.match(x)]
            bal = float(nums[-1].replace(",","")) if nums else None
            bal_disp = f"{bal:,.2f} {dr_cr}".strip() if bal is not None else None
            txs.append({"date":date,"description":"Balance B/F","cheque_no":"",
                        "vehicle":"","lpo_no":"","quantity":None,
                        "invoice_no":"","cu_invoice":"","location":"",
                        "currency":"KES","debit":0.0,"credit":0.0,
                        "balance":bal,"balance_disp":bal_disp,"is_bf":True})
            i+=1; continue

        # Normal Invoice line
        # Structure after date: Invoice VEHICLE [LPO_PREFIX] QTY INV_NO CU_INV LOCATION KES DEBIT CREDIT BALANCE DR
        desc = rest[0] if rest else "Invoice"
        rest = rest[1:]  # skip "Invoice"

        # Find QTY position (e.g. 245.000) — anchors the columns
        qty_idx = next((j for j,x in enumerate(rest) if QTY_N.match(x)), None)
        if qty_idx is None: i+=1; continue

        before_qty = rest[:qty_idx]    # VEHICLE tokens + optional LPO prefix
        after_qty  = rest[qty_idx:]    # QTY INV_NO CU_INV LOCATION KES DEBIT CREDIT BALANCE DR

        qty = float(after_qty[0].replace(",","")) if after_qty else None
        tail = list(after_qty[1:])

        # Strip DR/CR from end
        dr_cr = ""
        if tail and tail[-1] in ("DR","CR"): dr_cr = tail.pop()
        # Last 3 numbers: balance, credit, debit
        balance = float(tail.pop().replace(",","")) if tail and MONEY_N.match(tail[-1]) else None
        balance_disp = f"{balance:,.2f} {dr_cr}".strip() if balance is not None else None
        credit  = float(tail.pop().replace(",","")) if tail and MONEY_N.match(tail[-1]) else None
        debit   = float(tail.pop().replace(",","")) if tail and MONEY_N.match(tail[-1]) else None
        # Next: currency (KES)
        curr = tail.pop() if tail else "KES"
        # Remaining: inv_no, cu_inv, location (1-2 words)
        inv_no   = tail[0] if len(tail) >= 1 else ""
        cu_inv   = tail[1] if len(tail) >= 2 else ""
        location = " ".join(tail[2:]) if len(tail) >= 3 else ""

        # before_qty: vehicle plate + optional LPO prefix
        # LPO prefix: starts with POHH or is short numeric (4-5 digits like 5474)
        vehicle = ""; lpo_prefix = ""
        if before_qty:
            last = before_qty[-1]
            if re.match(r"^POHH\d+$", last) or re.match(r"^\d{4,5}$", last):
                lpo_prefix = last
                vehicle = " ".join(before_qty[:-1])
            else:
                vehicle = " ".join(before_qty)

        # Build full LPO number
        if lpo_prefix and lpo_suffix:
            lpo_no = lpo_prefix + lpo_suffix          # e.g. POHH0012606005663
        elif lpo_prefix:
            lpo_no = lpo_prefix                        # e.g. 5474 (Nakuru style)
        else:
            lpo_no = lpo_suffix                        # fallback

        txs.append({"date":date,"description":desc,"cheque_no":"",
                    "vehicle":vehicle,"lpo_no":lpo_no,"quantity":qty,
                    "invoice_no":inv_no,"cu_invoice":cu_inv,"location":location,
                    "currency":curr,"debit":debit,"credit":credit,"balance":balance,
                    "balance_disp":balance_disp,"is_bf":False})
        i += 1

    return hdr, txs, chqs

def excel_national(hdr,txs,chqs):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Statement"
    ws.sheet_properties.tabColor=C["national_blue"]
    for i,w in enumerate([12,12,14,20,18,10,14,26,12,6,16,16,18],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    banner(ws,1,13,"NATIONAL CEMENT COMPANY LTD  |  CUSTOMER STATEMENT",C["national_blue"])
    lf=Font(name="Arial",bold=True,size=9); vf=Font(name="Arial",size=9)
    lb=PatternFill("solid",start_color=C["info_bg"])
    al=Alignment(horizontal="left",vertical="center",indent=1)
    info=[("Customer",hdr.get("customer",""),"Supplier",hdr.get("supplier","")),
          ("Statement Date",hdr.get("statement_date",""),"PIN No",hdr.get("pin",""))]
    for r,(l1,v1,l2,v2) in enumerate(info,start=2):
        for ci,lbl,val in [(1,l1,v1),(8,l2,v2)]:
            ws.merge_cells(start_row=r,start_column=ci,end_row=r,end_column=ci+2)
            ws.merge_cells(start_row=r,start_column=ci+3,end_row=r,end_column=ci+5)
            lc=ws.cell(row=r,column=ci,value=lbl); lc.font=lf; lc.fill=lb; lc.alignment=al
            vc=ws.cell(row=r,column=ci+3,value=val); vc.font=vf; vc.alignment=al
        ws.row_dimensions[r].height=16
    H=4
    col_headers(ws,H,["Date","Description","Cheque No","Vehicle No.","LPO No","Qty",
                       "Invoice No.","CU Invoice No.","Location","Curr.",
                       "Debit (KES)","Credit (KES)","Balance (KES)"],C["national_blue"])
    D=H+1
    for ri,tx in enumerate(txs,start=D):
        bf="B/F" in str(tx.get("description",""))
        bg=C["info_bg"] if bf else (C["alt_row"] if (ri-D)%2==1 else None)
        vals=[tx["date"],tx["description"],tx["cheque_no"],tx["vehicle"],tx["lpo_no"],
              tx["quantity"],tx["invoice_no"],tx["cu_invoice"],tx["location"],tx["currency"],
              tx["debit"],tx["credit"],tx.get("balance_disp", tx["balance"])]
        aligns=["center","left","center","center","center","right","center","center",
                "center","center","right","right","right"]
        fmts=[None,None,None,None,None,"#,##0.000",None,None,None,None,MFMT,MFMT,None]
        fgs=["000000"]*10+[C["debit_fg"],C["credit_fg"],"000000"]
        for ci,(v,a,f,fg) in enumerate(zip(vals,aligns,fmts,fgs),1):
            wc(ws,ri,ci,v,align=a,fmt=f,fg=fg,bg=bg,bdr=thin("bottom"))
        ws.row_dimensions[ri].height=15
    T=D+len(txs)
    ws.merge_cells(f"A{T}:J{T}")
    wc(ws,T,1,"TOTAL",bold=True,fg=C["col_hdr_fg"],bg=C["national_blue"],align="right",bdr=thin(),ind=1)
    for ci,col in [(11,"K"),(12,"L")]:
        c=ws.cell(row=T,column=ci,value=f"=SUM({col}{D}:{col}{T-1})")
        c.font=Font(name="Arial",bold=True,size=9,color=C["col_hdr_fg"])
        c.fill=PatternFill("solid",start_color=C["national_blue"])
        c.number_format=MFMT; c.alignment=Alignment(horizontal="right",vertical="center"); c.border=thin()
    last_bal = txs[-1].get("balance_disp","") if txs else ""
    c=ws.cell(row=T,column=13,value=last_bal)
    c.font=Font(name="Arial",bold=True,size=9,color=C["col_hdr_fg"])
    c.fill=PatternFill("solid",start_color=C["national_blue"])
    c.alignment=Alignment(horizontal="right",vertical="center"); c.border=thin()
    ws.row_dimensions[T].height=18
    ws.freeze_panes=f"A{D}"; ws.auto_filter.ref=f"A{H}:M{T-1}"
    ws.page_setup.orientation="landscape"; ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1
    if chqs:
        ws2=wb.create_sheet("Cheques on Hand"); ws2.sheet_properties.tabColor="2E7D32"
        for col,w in [("A",16),("B",16),("C",18)]: ws2.column_dimensions[col].width=w
        ws2.merge_cells("A1:C1")
        wc(ws2,1,1,"CHEQUES ON HAND",bold=True,fg=C["col_hdr_fg"],bg="2E7D32",align="center",sz=12)
        ws2.row_dimensions[1].height=26
        for ci,h in enumerate(["Cheque No","Due On","Amount (KES)"],1):
            wc(ws2,2,ci,h,bold=True,fg=C["col_hdr_fg"],bg=C["national_blue"],align="center",bdr=thin())
        ws2.row_dimensions[2].height=18
        for ri,chq in enumerate(chqs,start=3):
            bg=C["alt_row"] if (ri-3)%2==1 else None
            wc(ws2,ri,1,chq["cheque_no"],align="center",bg=bg,bdr=thin("bottom"))
            wc(ws2,ri,2,chq["due_on"],align="center",bg=bg,bdr=thin("bottom"))
            wc(ws2,ri,3,chq["amount"],align="right",bg=bg,bdr=thin("bottom"),fmt=MFMT,fg=C["national_blue"])
            ws2.row_dimensions[ri].height=15
        CR=3+len(chqs)
        ws2.merge_cells(f"A{CR}:B{CR}")
        wc(ws2,CR,1,"TOTAL",bold=True,fg=C["col_hdr_fg"],bg=C["national_blue"],align="right",bdr=thin())
        c=ws2.cell(row=CR,column=3,value=f"=SUM(C3:C{CR-1})")
        c.font=Font(name="Arial",bold=True,size=9,color=C["col_hdr_fg"])
        c.fill=PatternFill("solid",start_color=C["national_blue"])
        c.number_format=MFMT; c.alignment=Alignment(horizontal="right",vertical="center"); c.border=thin()
        ws2.row_dimensions[CR].height=18
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ══════════════════════════════════════════════════════════════════════════════
# SAVANNAH CEMENT PARSER
# ══════════════════════════════════════════════════════════════════════════════
DR2=re.compile(r"^\d{2}\.\d{2}\.\d{2}$"); MR2=re.compile(r"^-?[\d,]+\.\d{2}$")
PFX=("IN ","CN ","RC ","DN ")

def parse_savannah(lines):
    hdr={"customer_name":"New Muthokinju Hardware Ltd","customer_code":"LC000072",
         "credit_limit":2_000_000.00,"sales_employee":"Fred Oguttu",
         "contact_person":"David","currency":"KES","printed_on":""}
    for ln in lines:
        m=re.search(r"Printed On:\s*\S+\s+(\S+)",ln)
        if m: hdr["printed_on"]=m.group(1)
    txs=[]; i=0
    while i<len(lines):
        ln=lines[i].strip()
        if not any(ln.startswith(p) for p in PFX): i+=1; continue
        p=ln.split(); dt=p[0]; dn=p[1] if len(p)>1 else ""; rest=p[2:]
        di=[j for j,x in enumerate(rest) if DR2.match(x)]
        mi=[j for j,x in enumerate(rest) if MR2.match(x)]
        if len(di)>=2:
            rn=" ".join(rest[:di[0]]); d=rest[di[0]]; dd=rest[di[1]]; af=di[1]+1
            if mi:
                fm=min(m for m in mi if m>di[1]); dps=rest[af:fm]; amts=rest[fm:]
            else: dps=rest[af:]; amts=[]
        else: rn=d=dd=""; dps=rest; amts=[]
        rex=[]; j=i+1
        while j<len(lines):
            nx=lines[j].strip()
            if any(nx.startswith(x) for x in PFX): break
            if nx.startswith("Page:") or nx.startswith("Total") or nx.startswith("Balance Due"): break
            if DR2.match(nx) or MR2.match(nx): break
            if any(k in nx for k in ("Savannah","Customer Statement","Currency:","Credit Limit:",
                "Sales Employee:","Contact Person:","Address:","Phone #:",
                "Document Ref.","Prior Period","Debit Credit")): break
            rex.append(nx); j+=1
        i=j
        dbt=crd=bal=None
        if len(amts)==2:
            if dt in ("CN","RC"): crd,bal=pn(amts[0]),pn(amts[1])
            else: dbt,bal=pn(amts[0]),pn(amts[1])
        elif len(amts)==3: dbt,crd,bal=pn(amts[0]),pn(amts[1]),pn(amts[2])
        elif len(amts)==1: bal=pn(amts[0])
        txs.append({"doc_type":dt,"doc_no":dn,"ref_no":(" ".join([rn]+rex)).strip(),
                    "date":d,"due_date":dd,"details":" ".join(dps).strip() or "A/R Invoices",
                    "debit":dbt,"credit":crd,"balance":bal})
    return hdr,txs

def excel_savannah(hdr,txs):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Statement"
    ws.sheet_properties.tabColor=C["savannah_orange"]
    for i,w in enumerate([6,8,30,11,11,32,16,16,18],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    banner(ws,1,9,"SAVANNAH CEMENT 2025 LIMITED  |  CUSTOMER STATEMENT",C["savannah_orange"])
    lf=Font(name="Arial",bold=True,size=9); vf=Font(name="Arial",size=9)
    lb=PatternFill("solid",start_color=C["info_bg"])
    al=Alignment(horizontal="left",vertical="center",indent=1)
    info=[("Customer Name",hdr["customer_name"],"Currency",hdr["currency"]),
          ("Customer Code",hdr["customer_code"],"Credit Limit",f"KES {hdr['credit_limit']:,.2f}"),
          ("Sales Employee",hdr["sales_employee"],"Contact",hdr["contact_person"]),
          ("Printed On",hdr["printed_on"],"","")]
    for r,(l1,v1,l2,v2) in enumerate(info,start=2):
        for ci,val,bold in [(1,l1,True),(3,v1,False),(6,l2,True),(8,v2,False)]:
            ws.merge_cells(start_row=r,start_column=ci,end_row=r,end_column=ci+1)
            c=ws.cell(row=r,column=ci,value=val)
            c.font=lf if bold else vf; c.alignment=al
            if bold: c.fill=lb
        ws.row_dimensions[r].height=16
    H=6
    col_headers(ws,H,["Type","Doc No.","Ref/Vehicle","Date","Due Date",
                       "Details","Debit (KES)","Credit (KES)","Balance (KES)"],C["savannah_orange"])
    D=H+1
    for ri,tx in enumerate(txs,start=D):
        cr=tx["doc_type"] in ("CN","RC")
        bg=C["credit_row"] if cr else (C["alt_row"] if (ri-D)%2==1 else None)
        vals=[tx["doc_type"],tx["doc_no"],tx["ref_no"],tx["date"],tx["due_date"],
              tx["details"],tx["debit"],tx["credit"],tx["balance"]]
        aligns=["center","center","left","center","center","left","right","right","right"]
        fmts=[None,None,None,None,None,None,MFMT,MFMT,MFMT]
        fgs=["000000","000000","000000","000000","000000","000000",
             C["debit_fg"],C["credit_fg"],"000000"]
        for ci,(v,a,f,fg) in enumerate(zip(vals,aligns,fmts,fgs),1):
            if ci==9 and v is not None: fg=C["neg_bal"] if v<0 else C["credit_fg"]
            wc(ws,ri,ci,v,align=a,fmt=f,fg=fg,bg=bg,bdr=thin("bottom"))
        ws.row_dimensions[ri].height=15
    T=D+len(txs)
    ws.merge_cells(f"A{T}:F{T}")
    wc(ws,T,1,"TOTAL",bold=True,fg=C["col_hdr_fg"],bg=C["header_dark"],align="right",bdr=thin(),ind=1)
    for ci,col in [(7,"G"),(8,"H"),(9,"I")]:
        c=ws.cell(row=T,column=ci,value=f"=SUM({col}{D}:{col}{T-1})")
        c.font=Font(name="Arial",bold=True,size=9,color=C["col_hdr_fg"])
        c.fill=PatternFill("solid",start_color=C["header_dark"])
        c.number_format=MFMT; c.alignment=Alignment(horizontal="right",vertical="center"); c.border=thin()
    ws.row_dimensions[T].height=18
    ws.freeze_panes=f"A{D}"; ws.auto_filter.ref=f"A{H}:I{T-1}"
    ws.page_setup.orientation="landscape"; ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ══════════════════════════════════════════════════════════════════════════════
# MAIN CONVERT
# ══════════════════════════════════════════════════════════════════════════════
def convert(pdf_bytes):
    lines=[]
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        pages=len(pdf.pages)
        for page in pdf.pages:
            t=page.extract_text()
            if t: lines.extend(t.split("\n"))
    fmt=detect(lines)
    if fmt=="mombasa":
        hdr,txs=parse_mombasa(lines)
        return fmt,len(txs),0,pages,excel_mombasa(hdr,txs)
    elif fmt=="karsan":
        hdr,txs=parse_karsan(lines)
        return fmt,len(txs),0,pages,excel_karsan(hdr,txs)
    elif fmt=="national":
        hdr,txs,chqs=parse_national(lines)
        return fmt,len(txs),len(chqs),pages,excel_national(hdr,txs,chqs)
    elif fmt=="savannah":
        hdr,txs=parse_savannah(lines)
        return fmt,len(txs),0,pages,excel_savannah(hdr,txs)
    return "unknown",0,0,pages,None

# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════
uploaded=st.file_uploader("Drop your PDF here",type=["pdf","PDF"],label_visibility="collapsed")

if uploaded:
    st.markdown(f"**Uploaded:** `{uploaded.name}` &nbsp;({uploaded.size/1024:.1f} KB)")
    with st.spinner("Converting…"):
        try:
            fmt,tx_count,chq_count,pages,buf=convert(uploaded.read())
        except Exception as e:
            st.error(f"❌ Error: {e}"); st.stop()

    if fmt=="unknown":
        st.warning("⚠️ Format not recognised. Supported: Savannah Cement, National Cement, Mombasa Cement, Karsan Ramji.")
    else:
        labels={"savannah":"Savannah Cement","national":"National Cement",
                "mombasa":"Mombasa Cement","karsan":"Karsan Ramji (Ndovu)"}
        label=labels.get(fmt,fmt)
        chq_txt=f" · {chq_count} cheques" if chq_count else ""
        st.markdown(f"""<div class="ok">✅ &nbsp;<b>{label}</b> converted &nbsp;·&nbsp;
        {pages} page(s) &nbsp;·&nbsp; {tx_count} transactions{chq_txt}</div>""",
        unsafe_allow_html=True)
        st.write("")
        st.download_button("⬇️  Download Excel File",data=buf,
            file_name=Path(uploaded.name).stem+"_converted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.markdown("---")
st.markdown("<div style='text-align:center;color:#94a3b8;font-size:.8rem'>"
            "Files processed in memory · never stored · free to use</div>",unsafe_allow_html=True)
